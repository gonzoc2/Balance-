import pandas as pd
import streamlit as st
from io import BytesIO
import requests
import locale
from openpyxl import load_workbook

# URL del archivo de mapeo
mapeo_url = 'https://docs.google.com/spreadsheets/d/1rt6Suyg1XgFxV0nTgblSkfZakzHrPNci/export?format=xlsx'
excel =  'https://docs.google.com/spreadsheets/d/1MevamCHCwCs0lvSJdbYmH1sCOqDeYpXu/export?format=xlsx'
base_wh = 'https://docs.google.com/spreadsheets/d/1yREufu125JBMsN1EE-5EXVZZGNeQ6pKs/export?format=xlsx'
@st.cache_data
def cargar_datos(url):
    response = requests.get(url)
    response.raise_for_status()  # Verifica si hubo error en la descarga
    archivo_excel = BytesIO(response.content)
    return pd.read_excel(archivo_excel, engine="openpyxl")

df_mapeo = cargar_datos(mapeo_url)
excel = cargar_datos(excel)
df_wh = cargar_datos(base_wh)
# Título de la app
st.title("Balance de comprobación ESGARI")
columnas_utiles = ['DEFAULT_EFFECTIVE_DATE', 'DESC_SEGMENT1', 'SEGMENT5', 'DEBIT', 'CREDIT']
df_wh = df_wh[columnas_utiles]
df_wh.columns = ['Fecha', 'Empresa', 'Cuenta', 'Débito', 'Crédito']
df = df_wh.copy()

# Optimización de tipos de datos
df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')  # Convertir a datetime
df['Cuenta'] = pd.to_numeric(df['Cuenta'], errors='coerce', downcast='integer')  # Convertir a numérico (entero)
df['Débito'] = pd.to_numeric(df['Débito'], errors='coerce', downcast='float')  # Convertir a flotante
df['Crédito'] = pd.to_numeric(df['Crédito'], errors='coerce', downcast='float')

# Extraer mes y año de manera eficiente
df['Año'] = df['Fecha'].dt.year
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
df['Mes'] = df['Fecha'].dt.strftime('%B')  # Convertir a nombre del mes en español

df.drop(columns=['Fecha'], inplace=True)

df = df.merge(df_mapeo, on='Cuenta', how='left')

df_sin_mapeo = df[df['nombre cuenta'].isna()]
df_sin_mapeo = df_sin_mapeo['Cuenta'].drop_duplicates()
if df_sin_mapeo.empty:  # Check df_sin_mapeo instead
    st.success('Todas las cuentas tienen mapeo')
else:
    st.warning(f'Cuentas sin mapeo: {len(df_sin_mapeo)}')
    st.dataframe(df_sin_mapeo)

# Filtrado basado en selección del usuario
col1, col2 = st.columns(2)
año = col1.selectbox('Selecciona un año', sorted(df['Año'].unique(), reverse=True))
mes = col2.selectbox('Selecciona un mes', df['Mes'].unique())
emp = st.selectbox('Selecciona una empresa', df['Empresa'].unique())

df_fecha = df[(df['Año'] == año) & (df['Mes'] == mes) & (df['Empresa'] == emp)]

# Agrupar con optimización
df_fecha = df_fecha.drop(columns=['Empresa'])
df_fecha = df_fecha.groupby(['Año', 'Mes', 'Cuenta', 'nombre cuenta', 'Categoria'], as_index=False).sum(numeric_only=True)
#datos antes de fecha elegida
meses_dict = {
"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
"julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
}
mes_num = meses_dict.get(mes)
df["Mes_Num"] = df["Mes"].map(meses_dict)
if mes_num == 1:  # Si es enero, solo tomar años anteriores
    df_saldos_iniciales = df[df["Año"] < año]
else:  # Si es otro mes, aplicar el filtro normal
    df_saldos_iniciales = df[(df["Año"] < año) | ((df["Año"] == año) & (df["Mes_Num"] < mes_num))]


df_saldos_iniciales = df_saldos_iniciales.drop(columns=['Empresa', 'Mes', 'Año'])
df_saldos_iniciales = df_saldos_iniciales.groupby(['Cuenta', 'nombre cuenta', 'Categoria'], as_index=False).sum(numeric_only=True)
df_saldos_iniciales['Saldo inicial'] = df_saldos_iniciales['Débito'] - df_saldos_iniciales['Crédito']
df_saldos_iniciales.drop(columns=['Débito', 'Crédito', 'Mes_Num'], inplace=True)
df_fecha = df_fecha.merge(df_saldos_iniciales, on=['Cuenta', 'nombre cuenta', 'Categoria'], how='outer')
df_fecha['Saldo inicial'] = df_fecha['Saldo inicial'].fillna(0)
# Rellenar NaN en débitos y créditos con 0
df_fecha[['Débito', 'Crédito']] = df_fecha[['Débito', 'Crédito']].fillna(0)

# Cálculo eficiente de 'Neto' usando numpy en lugar de apply
df_fecha['Año'] = df_fecha['Año'].fillna(año)
df_fecha['Mes'] = df_fecha['Mes'].fillna(mes)

orden = ['Año', 'Mes', 'Cuenta', 'nombre cuenta', 'Categoria', 'Saldo inicial', 'Débito', 'Crédito']
df_fecha = df_fecha[orden]
df_fecha['Saldo final'] = df_fecha['Saldo inicial'] + df_fecha['Débito'] - df_fecha['Crédito']
st.write(df_fecha)

# Botón para descargar Excel
def generar_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')
    return output.getvalue()

excel_data = generar_excel(df_fecha)

st.download_button(
    label="📥 Descargar Saldos",
    data=excel_data,
    file_name="saldos_cuentas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)        

ingreso_actual = df[(df['Año'] == año) & (df['Mes'] == mes) & (df['Empresa'] == emp)]
egreso_actual = df[(df['Año'] == año) & (df['Mes'] == mes) & (df['Empresa'] == emp)]
ingreso_actual[['Crédito', 'Débito']] = ingreso_actual[['Crédito', 'Débito']].fillna(0)
egreso_actual[['Crédito', 'Débito']] = egreso_actual[['Crédito', 'Débito']].fillna(0)

ingreso_actual = ingreso_actual[(ingreso_actual['Cuenta'] > 399999999) & (ingreso_actual['Cuenta']<500000000)]
ingreso_actual['Neto'] = ingreso_actual['Crédito'] - ingreso_actual['Débito']
ingreso_actual = ingreso_actual['Neto'].sum()
egreso_actual = egreso_actual[egreso_actual['Cuenta'] > 499999999]
egreso_actual['Neto'] = egreso_actual['Débito'] - egreso_actual['Crédito']
egreso_actual = egreso_actual['Neto'].sum()


df_fecha = df_fecha.drop(columns=['Año', 'Mes', 'Cuenta', 'Saldo inicial', 'Débito', 'Crédito', 'nombre cuenta'])
df_fecha = df_fecha.groupby(['Categoria'], as_index=False).sum(numeric_only=True)

fecha_r = f'{mes} de {año}'
fecha_esf = f'ESTADO DE SITUACION FIANCIERA {mes} {año}'
df_fecha.loc[len(df_fecha)] = ['nombre', emp]
df_fecha.loc[len(df_fecha)] = ['fecha_r', fecha_r]
df_fecha.loc[len(df_fecha)] = ['fecha', fecha_esf]

# Función para cargar datos desde una URL en un objeto BytesIO
@st.cache_data
def cargar_datos(url):
    response = requests.get(url)
    response.raise_for_status()  # Verifica si hubo error en la descarga
    return BytesIO(response.content)  # Devuelve el archivo en formato BytesIO

# URL del archivo original
excel_er = 'https://docs.google.com/spreadsheets/d/1yUqlBNTb4CM_ssWwNgktZ4Lx27IKEAOc/export?format=xlsx'
archivo_excel = cargar_datos(excel_er)  # Cargar archivo en memoria

# 📌 Cargar el archivo Excel en memoria y modificar solo "Hoja 1"
with BytesIO(archivo_excel.getvalue()) as excel_file:
    # Cargar el libro de trabajo de openpyxl
    book = load_workbook(excel_file)

    # Revisar si "Hoja 1" existe en el archivo
    if "Hoja1" in book.sheetnames:
        sheet = book["Hoja1"]  # Obtener la hoja
    else:
        sheet = book.create_sheet("Hoja1")  # Si no existe, crearla

    # Escribir `df_fecha` en la hoja existente sin eliminar el contenido anterior
    with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        writer.book = book  # Mantener todas las hojas
        df_fecha.to_excel(writer, sheet_name="Hoja1", index=False, startrow=sheet.max_row)

    # Guardar cambios en el archivo en memoria
    excel_file.seek(0)
    final_excel_bytes = excel_file.getvalue()

# Título en Streamlit
st.title(f'Balance de comprobación {emp}. {mes} de {año}')

# 📥 Botón de descarga en Streamlit
st.download_button(
    label="📥 Descargar balance",
    data=final_excel_bytes,
    file_name=f"balance_{mes}_{año}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)   


